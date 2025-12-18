import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
import streamlit_authenticator as stauth
import yaml
import os

# ==================== CONFIGURAÇÕES GERAIS E ESTILO ====================
st.set_page_config(
    page_title="Canivete Suíço Fiscal",
    page_icon="🔪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo customizado para deixar mais clean e profissional
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1e3a8a;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: 600;
    }
    .sub-header {
        font-size: 1.5rem;
        color: #1e40af;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .info-box {
        background-color: #f0f9ff;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #3b82f6;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #f0fdf4;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #22c55e;
    }
    .stButton>button {
        background-color: #3b82f6;
        color: white;
        border-radius: 8px;
        height: 3em;
        width: 100%;
        font-weight: 600;
    }
    .stDownloadButton>button {
        background-color: #10b981;
        color: white;
    }
    </style>
    """, unsafe_allow_html=True)

# Pastas e arquivos
DATA_DIR = "dados_clientes"
os.makedirs(DATA_DIR, exist_ok=True)
USERS_FILE = "users.yaml"

# Função para salvar config
def salvar_config(config_novo):
    with open(USERS_FILE, 'w', encoding='utf-8') as file:
        yaml.dump(config_novo, file, default_flow_style=False, allow_unicode=True)

# Carrega ou cria config
if not os.path.exists(USERS_FILE):
    config = {
        'credentials': {'usernames': {}},
        'cookie': {'name': 'canivete_cookie', 'key': 'chave_secreta_longa_2025', 'expiry_days': 30},
        'preauthorized': []
    }
    salvar_config(config)
else:
    with open(USERS_FILE, encoding='utf-8') as file:
        config = yaml.safe_load(file) or {
            'credentials': {'usernames': {}},
            'cookie': {'name': 'canivete_cookie', 'key': 'chave_secreta_longa_2025', 'expiry_days': 30},
            'preauthorized': []
        }

# Autenticador
authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days']
)

# Verifica se já tem contador
tem_contador = any(
    config['credentials']['usernames'].get(u, {}).get('role') == 'contador'
    for u in config['credentials']['usernames']
)

if not tem_contador:
    # ==================== TELA DE BOAS-VINDAS E CADASTRO ====================
    st.markdown("<h1 class='main-header'>🔪 Canivete Suíço Fiscal</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; font-size: 1.2rem; color: #64748b;'>A ferramenta que simplifica o envio de notas fiscais entre clientes e contadores</p>", unsafe_allow_html=True)
    
    st.markdown("### Configure sua conta de administrador")
    
    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        with st.form("cadastro_contador", clear_on_submit=True):
            st.markdown("#### Dados do contador principal")
            nome_contador = st.text_input("Nome completo ou do escritório")
            login_contador = st.text_input("Login de acesso", placeholder="ex: meuescritorio")
            senha_contador = st.text_input("Senha", type="password")
            confirma_senha = st.text_input("Confirme a senha", type="password")
            
            if st.form_submit_button("🚀 Criar conta e começar"):
                if not all([nome_contador, login_contador, senha_contador, confirma_senha]):
                    st.error("Preencha todos os campos.")
                elif senha_contador != confirma_senha:
                    st.error("As senhas não coincidem.")
                elif login_contador in config['credentials']['usernames']:
                    st.error("Este login já existe.")
                else:
                    hashed = stauth.Hasher([senha_contador]).generate()[0]
                    config['credentials']['usernames'][login_contador] = {
                        'name': nome_contador,
                        'password': hashed,
                        'role': 'contador'
                    }
                    salvar_config(config)
                    st.success("Conta criada com sucesso!")
                    st.balloons()
                    st.rerun()

else:
    # ==================== LOGIN ====================
    name, authentication_status, username = authenticator.login('Acesse sua conta', 'main')

    if authentication_status:
        authenticator.logout('Sair', 'sidebar')
        
        # Sidebar bonita
        with st.sidebar:
            st.markdown(f"### 👋 Olá, **{name}**!")
            st.markdown("---")
            st.markdown("**Canivete Suíço Fiscal**")
            st.markdown("Ferramenta exclusiva para contadores e clientes")
            st.markdown("---")
            st.caption("Versão 1.0 • 2025")

        user_data = config['credentials']['usernames'].get(username, {})
        role = user_data.get('role', 'cliente')

        # ==================== FUNÇÕES ====================
        def extrair_dados_nfe(arquivo):
            try:
                tree = ET.parse(arquivo)
                root = tree.getroot()
                ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                def get_text(path, default="-"):
                    elem = root.find(path, ns)
                    return elem.text.strip() if elem is not None and elem.text else default
                nota = get_text('.//nfe:ide/nfe:nNF', 'Não encontrada')
                data_str = get_text('.//nfe:ide/nfe:dhEmi', '')
                data = data_str[:10] if len(data_str) >= 10 else 'Não encontrada'
                cliente = get_text('.//nfe:dest/nfe:xNome', 'Consumidor Final')
                cnpj = get_text('.//nfe:dest/nfe:CNPJ', '')
                cpf = get_text('.//nfe:dest/nfe:CPF', '')
                documento = cnpj if cnpj != "-" else cpf if cpf != "-" else "Não informado"
                valor_total_str = get_text('.//nfe:total/nfe:ICMSTot/nfe:vNF', '0')
                valor_total = float(valor_total_str.replace(',', '.')) if valor_total_str != "-" else 0.0
                return {
                    "Número NF": nota,
                    "Data Emissão": data,
                    "Cliente": cliente,
                    "CNPJ/CPF": documento,
                    "Valor Total (R$)": valor_total,
                }
            except Exception:
                st.error(f"Erro ao processar {arquivo.name}")
                return None

        def gerar_excel_formatado(df):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Notas Fiscais', startrow=1)
                ws = writer.sheets['Notas Fiscais']
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 14
                ws.column_dimensions['C'].width = 40
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 18
                from openpyxl.styles import Font, Alignment, PatternFill
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
                for cell in ws[2]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for row in range(3, len(df) + 3):
                    ws.cell(row=row, column=2).number_format = 'DD/MM/YYYY'
                    ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
                total_row = len(df) + 3
                ws.cell(row=total_row, column=4, value="TOTAL FATURADO").font = Font(bold=True)
                ws.cell(row=total_row, column=5, value=df['Valor Total (R$)'].sum()).font = Font(bold=True)
                ws.cell(row=total_row, column=5).number_format = 'R$ #,##0.00'
                ws['A1'] = f"Relatório de Notas Fiscais - {datetime.now().strftime('%B/%Y')}"
                ws['A1'].font = Font(size=16, bold=True)
                ws.merge_cells('A1:E1')
            return output.getvalue()

        # ==================== PORTAL DO CLIENTE ====================
        if role == 'cliente':
            st.markdown("<h1 class='main-header'>Portal do Cliente</h1>", unsafe_allow_html=True)
            st.markdown(f"<p style='text-align: center; font-size: 1.2rem;'>Olá **{name}**! Envie suas notas fiscais do mês aqui.</p>", unsafe_allow_html=True)
            
            st.markdown("<div class='info-box'>"
                        "<strong>Dica:</strong> Arraste todos os XMLs do mês de uma vez. O sistema processa tudo automaticamente e gera a planilha pronta para seu contador."
                        "</div>", unsafe_allow_html=True)

            files = st.file_uploader(
                "Arraste seus arquivos XML aqui ou clique para selecionar",
                type=["xml"],
                accept_multiple_files=True,
                help="Arquivos XML de NF-e. Pode enviar quantos quiser."
            )

            if files:
                with st.spinner("Processando seus arquivos XML..."):
                    dados = [res for arq in files if (res := extrair_dados_nfe(arq))]
                    if dados:
                        df = pd.DataFrame(dados)
                        df['Data Emissão'] = pd.to_datetime(df['Data Emissão'], errors='coerce')
                        df = df.sort_values('Data Emissão', ascending=False)
                        df_display = df.copy()
                        df_display['Data Emissão'] = df_display['Data Emissão'].dt.strftime('%d/%m/%Y')
                        df_display['Valor Total (R$)'] = df_display['Valor Total (R$)'].map('R$ {:,.2f}'.format)

                        st.success(f"✅ {len(df)} notas processadas com sucesso!")
                        st.dataframe(df_display, use_container_width=True)

                        cliente_dir = os.path.join(DATA_DIR, username)
                        os.makedirs(cliente_dir, exist_ok=True)
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                        excel_data = gerar_excel_formatado(df)
                        nome_arquivo = f"relatorio_{timestamp}.xlsx"
                        caminho = os.path.join(cliente_dir, nome_arquivo)
                        with open(caminho, 'wb') as f:
                            f.write(excel_data)

                        st.download_button(
                            label="📥 Baixar planilha Excel pronta",
                            data=excel_data,
                            file_name=nome_arquivo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.markdown("<div class='success-box'>Seu contador já pode acessar esse relatório no painel dele.</div>", unsafe_allow_html=True)
                    else:
                        st.warning("Nenhum dado foi extraído dos arquivos enviados. Verifique se são XMLs de NF-e válidos.")

        # ==================== PAINEL DO CONTADOR ====================
        else:
            st.markdown("<h1 class='main-header'>Painel do Contador</h1>", unsafe_allow_html=True)
            tab1, tab2 = st.tabs(["📂 Relatórios Recebidos", "👥 Gerenciar Clientes"])

            with tab1:
                st.markdown("<h2 class='sub-header'>Relatórios enviados pelos clientes</h2>", unsafe_allow_html=True)
                clientes = [d for d in os.listdir(DATA_DIR) if os.path.isdir(os.path.join(DATA_DIR, d))]
                if not clientes:
                    st.info("Nenhum cliente enviou relatórios ainda.")
                else:
                    cliente_sel = st.selectbox("Selecione o cliente", sorted(clientes), key="relatorios_cliente")
                    pasta = os.path.join(DATA_DIR, cliente_sel)
                    arquivos = sorted([f for f in os.listdir(pasta) if f.endswith('.xlsx')], reverse=True)
                    if arquivos:
                        st.write(f"**Últimos relatórios de {cliente_sel}:**")
                        for arq in arquivos[:20]:
                            caminho = os.path.join(pasta, arq)
                            with open(caminho, 'rb') as f:
                                st.download_button(f"📥 {arq}", f.read(), arq, key=f"download_{arq}")
                    else:
                        st.info("Este cliente ainda não enviou relatórios.")

            with tab2:
                st.markdown("<h2 class='sub-header'>Gerenciar acessos de clientes</h2>", unsafe_allow_html=True)
                st.markdown("Crie um login para cada cliente que precisa enviar notas.")

                with st.expander("➕ Adicionar novo cliente", expanded=False):
                    with st.form("novo_cliente_form"):
                        col1, col2 = st.columns(2)
                        with col1:
                            nome_cli = st.text_input("Nome da empresa ou pessoa")
                            login_cli = st.text_input("Login do cliente", placeholder="ex: joao_silva")
                        with col2:
                            senha_cli = st.text_input("Senha temporária", type="password")
                            senha_confirm = st.text_input("Confirme a senha", type="password")
                        
                        if st.form_submit_button("Criar acesso"):
                            if not all([nome_cli, login_cli, senha_cli, senha_confirm]):
                                st.error("Preencha todos os campos.")
                            elif senha_cli != senha_confirm:
                                st.error("As senhas não coincidem.")
                            elif login_cli in config['credentials']['usernames']:
                                st.error("Este login já existe.")
                            else:
                                hashed = stauth.Hasher([senha_cli]).generate()[0]
                                config['credentials']['usernames'][login_cli] = {
                                    'name': nome_cli,
                                    'password': hashed,
                                    'role': 'cliente'
                                }
                                salvar_config(config)
                                st.success(f"Cliente **{nome_cli}** criado com sucesso!")
                                st.info(f"**Dados de acesso:**\n\nLogin: `{login_cli}`\nSenha: `{senha_cli}`")

                st.markdown("<h3 style='margin-top: 3rem;'>Clientes com acesso</h3>", unsafe_allow_html=True)
                clientes_cadastrados = [
                    u for u in config['credentials']['usernames']
                    if config['credentials']['usernames'][u].get('role') == 'cliente'
                ]
                if clientes_cadastrados:
                    for u in clientes_cadastrados:
                        col1, col2, col3 = st.columns([3,2,1])
                        col1.write(f"**{config['credentials']['usernames'][u]['name']}**")
                        col2.write(f"`{u}`")
                        if col3.button("Excluir", key=f"del_{u}"):
                            del config['credentials']['usernames'][u]
                            salvar_config(config)
                            st.rerun()
                else:
                    st.info("Nenhum cliente cadastrado ainda.")

    elif authentication_status is False:
        st.error("Usuário ou senha incorretos.")
    elif authentication_status is None:
        st.warning("Preencha os campos de login acima.")

st.caption("Canivete Suíço Fiscal • Desenvolvido para facilitar a rotina contábil • 2025")